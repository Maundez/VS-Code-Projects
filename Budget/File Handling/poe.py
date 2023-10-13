import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Select the active sheet
sheet = workbook.active

# Define the arrays for categories
animal_categories = ['dog', 'cat', 'horse']
greeting_categories = ['hello', 'goodbye']

# Iterate through each row in the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    value = row[0]  # Assuming category is in column 1 (0-based index)

    if value in animal_categories:
        sheet.cell(row=row[0].row, column=2).value = 'animal'  # Write 'animal' in column 2
    elif value in greeting_categories:
        sheet.cell(row=sheet.row_dimensions[row[0].row].index, column=2).value = 'greeting'  # Write 'greeting' in column 2

# Save the modified workbook
workbook.save('example.xlsx')
