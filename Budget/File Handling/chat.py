import openpyxl

# Load the original workbook
workbook = openpyxl.load_workbook('visa.xlsx')  # No need for data_only=True

# Access the 'lookup' worksheet
sheet = workbook['lookup']

# Define the new arrays for categories
animal_categories = [
    "ADDITIONALCARDF",
    "AHPCPHARMACYOUT",
    "ALDISTORES-NORT",
    "AMAZONAUSYDNEYS",
    "AmazonDigitalSv",
    "AMAZONMARKETPLA",
    "ANNUALFEE",
    "BAKERSDELIGHTGR",
    "Beast&CoSurryHi",
    "BlowBar+6113002",
    "BulkNutrients61"
]

# Initialize a variable to start from the first row
current_row = 1  # Start from the first row

# Trigger recalculation (optional)
sheet.calculate_dimension()

# Iterate through each cell in column 1 (column A) of the 'lookup' worksheet
for cell in sheet.iter_cols(min_col=1, max_col=1, values_only=True):
    value = cell[0]  # Get the value from the cell in column 1

    # Copy the value from column 0 to column 1 as plain text
    #sheet.cell(row=current_row, column=1).value = str(value)

    if value in animal_categories:
        sheet.cell(row=current_row, column=3).value = 'animal'  # Write 'animal' in column 3
    #else:
        #sheet.cell(row=current_row, column=3).value = ''  # Clear column 2 if not an animal category

    current_row += 1  # Increment the row number

# Save the modified workbook with a different name
workbook.save('modified_visa.xlsx')







