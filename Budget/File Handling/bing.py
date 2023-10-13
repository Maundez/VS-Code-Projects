import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Get the active sheet
sheet = workbook.active

# Print the sheet name
print(sheet.title)

# Read the cell values
for row in sheet.rows:
  for cell in row:
    print(cell.value)

# Iterate over all cells in the sheet
for row in sheet.iter_rows():
    for cell in row:
        # Check if the cell contains 'hello' or 'goodbye'
        if 'hello' in str(cell.value).lower() or 'goodbye' in str(cell.value).lower():
            # Write 'greeting' in the next column
            sheet.cell(row=cell.row, column=cell.column+1).value = 'greeting'
           
        # Check if the cell contains 'cat', 'dog', or 'elephant'
        elif 'cat' in str(cell.value).lower() or 'dog' in str(cell.value).lower() or 'elephant' in str(cell.value).lower():
            # Write 'animal' in the next column
            sheet.cell(row=cell.row, column=cell.column+1).value = 'animal'
            

# Save the changes to the workbook
workbook.save('example.xlsx')